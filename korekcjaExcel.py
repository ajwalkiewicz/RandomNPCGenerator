# Korekcja pliku excel
import openpyxl
import os
# ustawieni sciezki dla plików
sciezka = os.path.join('C:\\', 'Users', 'ajwal', 'OneDrive', 'IT', 'Projekty', 'Zew Cthulhu')
os.chdir(sciezka)
# otworzenie excela
workBook = openpyxl.load_workbook('lista_imion.xlsx')


def zmiana(workbook, column):
    sheet = workBook[workbook]
    for i in range(1, len(sheet[column])):
        sheet.cell(row=i, column=2).value = sheet.cell(
            row=i, column=2).value[0].upper() + sheet.cell(row=i, column=2).value[1:].lower()


zmiana('last names', 'B')
workBook.save('lista_imion_copy.xlsx')
