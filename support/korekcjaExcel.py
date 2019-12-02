'''
korekcjaExcel.py - zmienia nazwiska zapisane w pliku lista_imion.xlsx
z imion pisanych capslockiem na na imiona pisane z dużej litery
Wyniki zapisywanie są w pliku lista_imion_copy.xlsx
'''
import openpyxl

# otworzenie excela
excel_file = openpyxl.load_workbook('lista_imion.xlsx')


def zmiana(workbook, column):
    sheet = excel_file[workbook]
    for i in range(1, len(sheet[column])):
        sheet.cell(row=i, column=2).value = sheet.cell(
            row=i, column=2).value[0].upper() + sheet.cell(row=i, column=2).value[1:].lower()


# def usun_spacje(workbook):
#     sheet = excel_file[workbook]
#     for i in range(1, 11):
#         for j in range(2, 29):
#             if str(sheet.cell(row=j, column=i).value)[0] == ' ':
#                 sheet.cell(row=j, column=i).value = str(sheet.cell(row=j, column=i).value)[1:]
#
#
# def zmiana2(workbook, column):
#     sheet = excel_file[workbook]
#     for i in range(1, len(sheet[column])):
#         if sheet.cell(row=i, column=1).value[-1] == ' ':
#             sheet.cell(row=i, column=1).value = sheet.cell(row=i, column=1).value[0:-1]


zmiana('last names', 'B')
# usun_spacje('occupations')
# zmiana2('skills', 'A')
excel_file.save('data_copy.xlsx')
